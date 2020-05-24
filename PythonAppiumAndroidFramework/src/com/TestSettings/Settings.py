'''initial settings'''

import logging
import os

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.workbook.workbook import Worksheet

from src.com.FunctionLibrary.webdriver_factory import WebdriverFactory
from src.com.TestPlanInfo.TestCaseResult import TestCaseResult
from src.com.TestPlanInfo.TestPlanResult import TestPlanResult
from src.com.Utility.Tool import Tool


class Settings(object):
    """description of class"""
    testPlanResult: TestPlanResult = None
    CurrentScript: str = None
    CurrentScenario: str = None
    CurrentScriptPath: str = None
    CurrentScriptResultFolder: str = None
    StateFile: str = None
    StateFilePath: str = None
    TestCaseNumbers: int = 0
    isRunable: bool = True
    planFile: str = None
    planFilePath: str = None
    CurrentTestResultFolder: str = None
    wedriver = None
    GetParameters: dict = {}
    StoredParameters: dict = {}
    wdf = None
    desired_caps = {"platformName": "Android",
                    "platformVersion": "7.1.2",
                    'deviceName': 'Redmi 4x',
                    'udid': 'f890d0db7ce4',
                    'appPackage': 'com.miui.calculator',
                    'appActivity': 'com.miui.calculator.cal.CalculatorActivity',
                    'unicodeKeyboard': True,
                    'resetunicodeKeyboard': True,
                    'noReset': True,
                    'deviceReadyTimeout': 120,
                    'newCommandTimeout': 20,
                    'androidDeviceReadyTimeout': 60,
                    'autoAcceptAlerts': True,
                    'autoGrantPermissions': True
                    }

    def __init__(self):
        self.isRunable: bool = True

    @staticmethod
    def constructSettings():
        Settings._settings = Settings()
        Settings.planFile = 'Plan1'
        currentWD = os.getcwd()
        Settings.planFilePath = currentWD + '\\TestPlan\\Plan1.xlsx'
        tool = Tool()
        folderName: str = tool.RanString
        Settings.CurrentTestResultFolder = currentWD + \
            '\\TestResult\\' + folderName + '\\' + Settings.planFile
        os.makedirs(Settings.CurrentTestResultFolder)
        Settings.testPlanResult = TestPlanResult("TestPlan")
        LOG_FORMAT = "%(asctime)s - %(levelname)s - %(message)s"
        logging.basicConfig(filename=os.path.join(currentWD, 'TestResult', folderName,
                                                  'runtime_{0}.log'.format(folderName)),
                            level=logging.INFO, format=LOG_FORMAT)

    @staticmethod
    def getSettings():
        return Settings._settings

    def createWebDriver(self, command_executor):
        self.wdf = WebdriverFactory()
        self.wedriver = self.wdf.create(command_executor, self.desired_caps)

    def cleanUpAndSetUpPlanFile(self, planSheet: Worksheet, rowNum: int):
        planSheet.cell(rowNum, 3).value = "NotRun"
        planSheet.cell(rowNum, 4).value = "NotRun"
        planSheet.cell(rowNum, 5).value = "NotRun"
        self.CurrentScenario = planSheet.cell(rowNum, 1).value
        self.CurrentScript = self.CurrentScenario.replace('.xlsx', '')
        self.CurrentScriptResultFolder = self.CurrentTestResultFolder + '/' + self.CurrentScript
        os.makedirs(self.CurrentScriptResultFolder)
        self.CurrentScriptPath = os.getcwd() + '\\TestScript\\' + self.CurrentScenario
        testCaseResult = TestCaseResult(self.CurrentScenario)
        testCaseResult.TestCaseNumber = str(rowNum - 1)
        return testCaseResult

    def loadPanFileAndSetCaseNumber(self):
        wb_plan = load_workbook(self.planFilePath)
        logging.info(
            'open plan file: {0} successfully'.format(self.planFilePath))
        wk_plan_sheet = wb_plan['Sheet1']
        self.TestCaseNumbers = wk_plan_sheet.max_row - 1
        return wb_plan

    def loadScriptFile(self):
        wb_state = load_workbook(self.CurrentScriptPath)
        logging.info('open script file: "{0}" successfully'.format(
            self.CurrentScriptPath))
        return wb_state

    def getStateNumber(self, workB: Workbook):
        wk_state_sheet = workB['Sheet1']
        return wk_state_sheet.max_row - 1

    def setStateFile(self, stateSheet: Worksheet, rowNum: int):
        self.StateFile = stateSheet.cell(rowNum, 3).value
        self.StateFilePath = self.CurrentScriptPath.split(
            '.xlsx')[0] + '/' + self.StateFile

    def updateStateResult(self, stateSheet: Worksheet, testStateResult, rowNum: int):
        testStateResult.Attachment = self.StateFile
        testStateResult.TestStateDescription = stateSheet.cell(
            rowNum, 1).value if stateSheet.cell(rowNum, 1).value != None else ''
        testStateResult.TestStateNumber = str(rowNum - 1)
        # update state file
        stateSheet.cell(rowNum, 7).value = testStateResult.Attachment
        stateSheet.cell(rowNum, 8).value = testStateResult.StateResult
        stateSheet.cell(rowNum, 9).value = testStateResult.StateLog

        return testStateResult

    def updateScriptFileAndCaseResult(self, workB: Workbook,
                                      testCaseResult: TestCaseResult, caseResult):
        workB.save(self.CurrentScriptPath)
        logging.info('save scenario file: "{0}" successfully'.format(
            self.CurrentScriptPath))
        testCaseResult.CaseResult = caseResult
        testCaseResult.EndTime = Tool.CurrentTime()
        self.testPlanResult.TestCaseResultCollection.append(testCaseResult)

    def updatePlanFile(self, planSheet: Worksheet, rowNum):
        planSheet.cell(
            rowNum, 3).value = self.testPlanResult.TestCaseResultCollection[rowNum - 2].StartTime
        planSheet.cell(
            rowNum, 4).value = self.testPlanResult.TestCaseResultCollection[rowNum - 2].EndTime
        planSheet.cell(
            rowNum, 5).value = self.testPlanResult.TestCaseResultCollection[rowNum - 2].CaseResult

    @staticmethod
    def Get_InParameter(parameter: str):
        if(parameter == None or parameter == '-'):
            return None
        parameters = parameter.split('\n')
        for para in parameters:
            param: list = para.split('=')
            if(param.count == 1):
                raise Exception("Data format issue")
            else:
                paraName = param[0].strip()
                parValue = None
                if(param.count != 2):
                    paraValue = para.split(param[0] + '=')[1].strip()
                else:
                    parValue = param[0].strip()
                if (param[1].find("!") == -1):
                    if (param[1].find("AS") == -1):
                        paraValue = Settings.convertToStringList(paraValue)
                        Settings.GetParameters[paraName] = paraValue
                    else:
                        paraValue = param[1].split("AS")[0].strip()
                        paraValue = Settings.convertToStringList(paraValue)
                        Settings.GetParameters[paraName] = paraValue
                        storedParaName = param[1].split("AS")[1].strip()
                        Settings.StoredParameters[storedParaName] = paraValue
                else:
                    paraValue = Settings.StoredParameters[param[1].split('!')[
                        1].strip()]
                    Settings.GetParameters[paraName] = parValue
        return Settings.GetParameters

    @staticmethod
    def convertToStringList(paraValue: str):
        if (paraValue.startswith("[") and paraValue.endsWith("]")):
            paraValue = paraValue.split('[')[1].split(']')[0]
            convertedStringList = paraValue.split('|')
            return convertedStringList
        else:
            return paraValue
