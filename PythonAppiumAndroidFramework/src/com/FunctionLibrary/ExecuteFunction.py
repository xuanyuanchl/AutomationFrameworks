import logging
import traceback

from openpyxl import load_workbook

from src.com.Enumeration.StepResult import StepResult
from src.com.FunctionLibrary.LowLevelKeyword import LowLevelKeyword
from src.com.ObjectRepository.PageObject import PageObject
from src.com.TestPlanInfo.TestCaseResult import TestCaseResult
from src.com.TestPlanInfo.TestStateResult import TestStateResult
from src.com.TestPlanInfo.TestStepResult import TestStepResult
from src.com.TestSettings.Settings import Settings
from src.com.Utility.Tool import Tool


class ExecuteFunction(object):
    """execute the low/high level function"""

    def __init__(self):
        self.settings = Settings.getSettings()
        self.pageObject = PageObject()

    def _deco(self, func, keyword,  pageObjectDict, parameter):
        if self.settings.isRunable == False:
            return self.setNon_ExecutedResult(keyword)
        else:
            self.paras = self.settings.Get_InParameter(parameter)
            try:
                return func(keyword, pageObjectDict, parameter)
            except Exception as e:
                return self.setErrorResult(e)

    # args[0] is path expression
    # args[1] is selector like id, css selector, xpath
    # args[2] is the text need to input
    def keyword_Execute(self, keyword, pageObjectDict, parameter):
        llk = LowLevelKeyword()
        if keyword == llk.launchApp.__name__:
            return llk._deco(llk.launchApp, self.paras["command_executor"])
        elif keyword == llk.click.__name__:
            return llk._deco(llk.click, pageObjectDict)
        ######To be added other keywords####
        else:
            tool = Tool()
            testStepResult: TestStateResult = TestStepResult()
            testStepResult.endTime = Tool.CurrentTime()
            testStepResult.stepLog = str(
                testStepResult.endTime) + ' No such kind of keyword: ' + keyword
            testStepResult.stepResult = StepResult.Error
            self.settings.isRunable = False
            fileName = tool.RanString
            Tool.saveScreenShot(
                self.settings.wedriver, self.settings.CurrentScriptResultFolder + '\\' + fileName + '.png')
            testStepResult.Actual = fileName + '.png'
            return testStepResult

    # Read test state file, execute each test step
    def SetState(self, stateFile):
        self.settings = Settings.getSettings()
        self.testStateResult = TestStateResult(stateFile)
        try:
            wb_state = load_workbook(self.settings.StateFilePath)
            logging.info('open state file: "{0}" successfully'.format(
                self.settings.StateFilePath))
            wk_state_sheet = wb_state['Sheet1']
            testScriptRow: int = wk_state_sheet.max_row - 1
            # Clean last time result
            for i in range(2, testScriptRow + 2):
                wk_state_sheet.cell(i, 7).value = ''
                wk_state_sheet.cell(i, 8).value = 'Not run'
                wk_state_sheet.cell(i, 9).value = ''
            for i in range(2, testScriptRow + 2):
                description = wk_state_sheet.cell(
                    i, 1).value if wk_state_sheet.cell(i, 1).value != None else ''
                keyword = wk_state_sheet.cell(i, 2).value if wk_state_sheet.cell(
                    i, 2).value != None else ''
                parameter = wk_state_sheet.cell(
                    i, 3).value if wk_state_sheet.cell(i, 3).value != None else ''
                objectName = wk_state_sheet.cell(
                    i, 4).value if wk_state_sheet.cell(i, 4).value != None else ''
                tagName = wk_state_sheet.cell(i, 5).value if wk_state_sheet.cell(
                    i, 5).value != None else ''
                expect = wk_state_sheet.cell(i, 6).value if wk_state_sheet.cell(
                    i, 6).value != None else ''
                # execute test step
                obj = None
                if (objectName != None and objectName != '-' and objectName != ''):
                    obj = self.pageObject.getPageObject(objectName)
                if (isinstance(obj, dict)):
                    tagName = list(obj.keys())[0]
                testStepResult: TestStepResult = self._deco(
                    self.keyword_Execute, keyword, obj, parameter)
                logging.info(f'execute "{keyword}" ')
                testStepResult.StepNumber = str(i - 1)
                testStepResult.StepDescription = description
                testStepResult.StepKeyword = keyword
                testStepResult.Expect = '../../../../..' + self.settings.CurrentScriptPath.replace(
                    '.xlsx', '').split('src')[1] + "/Attachment/" + expect
                testStepResult.StepTagName = tagName
                testStepResult.StepParameter = parameter
                testStepResult.StepObject = objectName

                if testStepResult.Actual == None:
                    testStepResult.Actual = ''
                wk_state_sheet.cell(i, 5).value = tagName
                wk_state_sheet.cell(i, 7).value = testStepResult.Actual
                wk_state_sheet.cell(
                    i, 8).value = testStepResult.stepResult.name
                wk_state_sheet.cell(i, 9).value = testStepResult.stepLog
                self.testStateResult.TestStepResultCollection.append(
                    testStepResult)
            # save module file
            self.testStateResult.StateResult = self.TestState_Result(
                self.testStateResult)
            self.testStateResult.setLog()
            return self.testStateResult
        except Exception as e:
            logging.error('there is error: ' + str(e) +
                          '.' + traceback.format_exc())
        finally:
            wb_state.save(self.settings.StateFilePath)
            logging.info('save state file: {0} successfully'.format(
                self.settings.StateFilePath))

    # Update test state result
    @staticmethod
    def TestState_Result(testStateResult: TestStateResult):
        error_count = 0
        passed_count = 0
        failed_count = 0
        notrun_count = 0
        for testStepR in testStateResult.TestStepResultCollection:
            if testStepR.stepResult == StepResult.Error:
                error_count = error_count + 1
            elif testStepR.stepResult == StepResult.Failed:
                failed_count = failed_count + 1
            elif testStepR.stepResult == StepResult.Passed:
                passed_count = passed_count + 1
            else:
                notrun_count = notrun_count + 1
        if error_count != 0:
            return "Error"
        elif (error_count == 0 and failed_count == 0 and notrun_count == 0):
            return "Passed"
        elif(error_count == 0 and failed_count == 0 and passed_count == 0):
            return "NotRun"
        else:
            return "Failed"

    # Update test case result
    @staticmethod
    def TestCase_Result(testCaseResult: TestCaseResult):
        error_count = 0
        passed_count = 0
        failed_count = 0
        notrun_count = 0
        for testStateR in testCaseResult.TestStateResultCollection:
            if testStateR.StateResult == 'Error':
                error_count = error_count + 1
            elif testStateR.StateResult == 'Failed':
                failed_count = failed_count + 1
            elif testStateR.StateResult == 'Passed':
                passed_count = passed_count + 1
            else:
                notrun_count = notrun_count + 1
        if error_count != 0:
            return "Error"
        elif (error_count == 0 and failed_count == 0 and notrun_count == 0):
            return "Passed"
        elif(error_count == 0 and failed_count == 0 and passed_count == 0):
            return "NotRun"
        else:
            return "Failed"

    def setNon_ExecutedResult(self, methodName):
        testStepResult: TestStateResult = TestStepResult()
        testStepResult.startTime = Tool.CurrentTime()
        testStepResult.stepResult = StepResult.NotRun
        testStepResult.endTime = Tool.CurrentTime()
        testStepResult.stepLog = methodName + \
            " execution was " + testStepResult.stepResult.name
        return testStepResult

    def setErrorResult(self, e: Exception):
        tool = Tool()
        self.settings._isRunable = False
        testStepResult: TestStepResult = TestStepResult()
        testStepResult.startTime = tool.CurrentTime()
        testStepResult.stepResult = StepResult.Error
        testStepResult.endTime = tool.CurrentTime()
        testStepResult.stepLog = testStepResult.endTime + " " + str(e)
        fileName = tool.RanString
        Tool.saveScreenShot(
            self.settings.wedriver, self.settings.CurrentScriptResultFolder + '\\' + fileName + '.png')
        testStepResult.Actual = fileName + '.png'
        return testStepResult
