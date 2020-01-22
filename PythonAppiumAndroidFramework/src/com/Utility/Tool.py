import datetime

from appium import webdriver
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


class Tool(object):
    """description of class"""
    @property
    def RanString(self):
        return datetime.datetime.strftime(datetime.datetime.now(), '%m%d_%H%M%S')

    @staticmethod
    def CurrentTime():
        return datetime.datetime.strftime(datetime.datetime.now(), '%m-%d %H:%M:%S')

    @staticmethod
    def closeBrowser(wedriver: webdriver):
        wedriver.close()

    @staticmethod
    def readFile(fileName):
        with open(fileName, 'r') as fb:
            content = fb.read()
            return content

    @staticmethod
    def writeFile(fileName, text):
        with open(fileName, 'w') as fb:
            fb.write(text)

    @staticmethod
    def appendFile(fileName, text):
        with open(fileName, 'a') as fb:
            fb.write(text)

    @staticmethod
    def createFile(fiename):
        with open(fiename, 'w'):
            '''create a file'''

    @staticmethod
    def saveScreenShot(wedriver: webdriver, fileName):
        wedriver.get_screenshot_as_file(fileName)

    @staticmethod
    def openExcel(excelName):
        wb = load_workbook(excelName)
        return wb

    @staticmethod
    def getMaxRow(wb: Workbook, sheetName: str):
        wb_sheet = wb[sheetName]
        return wb_sheet.max_row
