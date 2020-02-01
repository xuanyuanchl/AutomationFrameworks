from builtins import staticmethod
import datetime
import json

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from selenium import webdriver


class Tool(object):
    """description of class"""
    @property
    def RanString(self):
        return datetime.datetime.strftime(datetime.datetime.now(), '%m%d_%H%M%S')

    @staticmethod
    def CurrentTime():
        return datetime.datetime.strftime(datetime.datetime.now(), '%m-%d %H:%M:%S')

    @staticmethod
    def closeBrowser(wedriver: webdriver):
        wedriver.close()

    @staticmethod
    def readFile(fileName):
        with open(fileName, 'r') as fb:
            content = fb.read()
            return content

    @staticmethod
    def writeFile(fileName, text):
        with open(fileName, 'w') as fb:
            fb.write(text)

    @staticmethod
    def appendFile(fileName, text):
        with open(fileName, 'a') as fb:
            fb.write(text)

    @staticmethod
    def createFile(fiename):
        with open(fiename, 'w'):
            '''create a file'''

    @staticmethod
    def saveScreenShot(wedriver: webdriver, fileName):
        wedriver.save_screenshot(fileName)

    @staticmethod
    def openExcel(excelName):
        wb = load_workbook(excelName)
        return wb

    @staticmethod
    def getMaxRow(wb: Workbook, sheetName: str):
        wb_sheet = wb[sheetName]
        return wb_sheet.max_row


class jsonParse(object):
    data: object = None

    def loadJson(self, file):
        with open(file) as fp:
            self.data = json.load(fp)
            return self.data

    def getValue(self, key):
        return self.data[key]
